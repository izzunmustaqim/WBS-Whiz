"""File parsing and validation for WBS Enhancement Application.

Pure functions for reading, validating, and parsing Excel files.
No GUI dependencies — callers are responsible for error display.
"""

import json
import os
import re

import pandas as pd
from openpyxl import load_workbook


def read_excel_file(
    file: str, rowskip: int = 0, rangecol: list | None = None
) -> pd.DataFrame:
    """Read an Excel file with pandas, clean and return a DataFrame.

    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file exceeds 25MB.
        pd.errors.EmptyDataError: If the Excel file has no data.
        pd.errors.ParserError: If the file can't be parsed.
    """
    file_size = os.path.getsize(file)
    max_mb = 25 * 1024 * 1024  # 25MB in bytes
    if file_size > max_mb:
        raise ValueError(
            "The member skillset file is too large. "
            "Please upload a file smaller than 25MB and try again"
        )

    file_data = pd.read_excel(file, skiprows=rowskip, usecols=rangecol)

    # Check if the file is empty
    if file_data.empty:
        raise pd.errors.EmptyDataError("The file is empty")

    # Data cleaning steps
    file_data.dropna(axis=1, how='all', inplace=True)
    file_data.fillna(0, inplace=True)

    return file_data


def check_file_validity(
    sheet_names: list[str], workbook, file_path: str
) -> tuple[bool, str]:
    """Validate an Excel workbook's structure and content.

    Returns:
        Tuple of (is_valid, message). If is_valid is False, message
        contains the error description.
    """
    file_size = os.path.getsize(file_path)
    max_mb = 25 * 1024 * 1024  # 25MB in bytes
    if file_size > max_mb:
        return (False, "Error! The SS file is too large. Please upload a file smaller than 25MB")

    for sheet_name in sheet_names:
        sheet_temp = workbook[sheet_name]
        if (sheet_temp.max_row == 1 and sheet_temp.max_column == 1
                and sheet_temp.cell(row=1, column=1).value is None):
            return (False, "Error! The screen layout files are empty. Please upload the correct file")

        for row in sheet_temp.iter_rows(values_only=True):
            filtered_row = [cell for cell in row if cell is not None]
            if filtered_row == []:
                return (False, "Error! The screen layout files are empty. Please upload the correct file and start the process again")

            elif "Screen Layout" in file_path and '画面レイアウト\n/Screen Layout' not in filtered_row:
                return (False, "Error! The file is not a Screen Layout file. Please upload the correct file and start the process again")

            elif "Application Detailed Specification" in file_path and "アプリケーション詳細仕様\n/Application Detailed Specification" not in filtered_row:
                return (False, "Error! The file is not an Application Detailed Specification file. Please upload the correct file and start the process again")

            elif "Event Process Sequence Diagram History" in file_path and "イベント処理シーケンス図\n/Event Process Sequence Diagram" not in filtered_row:
                return (False, "Error! The file is not an Event Process Sequence Diagram History file. Please upload the correct file and start the process again")

            else:
                return (True, "File is valid")

    return (True, "File is valid")


def extract_screen_name(file_path: str) -> str:
    """Extract screen/component name from a file path using regex.

    Looks for pattern: \\<name>_ in the file path and returns <name>.
    If the captured part contains backslashes, returns the last segment.
    """
    match = re.search(r"\\([^_]+)_", file_path)
    if match:
        extracted_part = match.group(1)
        backslash_count = extracted_part.count('\\')
        if backslash_count > 0:
            parts = extracted_part.split('\\')
            return parts[-1]
        else:
            return extracted_part
    return ""


def parse_screen_layout(
    file: str, sheet_name: str, keywords_header: list[str]
) -> tuple[str, str]:
    """Parse a Screen Layout Excel file and return JSON + screen name.

    Returns:
        Tuple of (json_string, screen_name).

    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file is empty or invalid.
        Exception: For any other parsing errors.
    """
    workbook = load_workbook(filename=file)

    sheet = workbook[sheet_name]
    screen_layout_data = []
    start_found = False

    for row in sheet.iter_rows(values_only=True):
        filtered_row = []
        for cell in row:
            if cell is not None:
                filtered_row.append(cell)

        if any(keyword in str(cell) for keyword in keywords_header for cell in filtered_row):
            start_found = True

        if start_found:
            if filtered_row != [] and '画面項目名\n/Screen Item Name' not in filtered_row:
                screen_layout_data.append(filtered_row)

    screen_name = extract_screen_name(file) + "_UI"

    # Initialize the JSON structure
    screen_layout_json = {
        screen_name: [],
    }

    for row in screen_layout_data:
        if len(row) > 1 and (row[1] != '-' or row[2] != '-'):
            screen_layout_json[screen_name].append({
                "Screen Item Name": row[1],
                "Type": row[2],
            })

    # Convert to JSON string for readability
    json_string = json.dumps(screen_layout_json, indent=4)
    return (json_string, screen_name)


def parse_app_detailed_spec(file_path: str) -> list:
    """Parse an Application Detailed Specification Excel file.

    Returns:
        List of filtered rows from the spec sheet.

    Raises:
        FileNotFoundError: If the file doesn't exist.
        Exception: For any other parsing errors.
    """
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet_names = [sheet.title for sheet in workbook.worksheets]

    sheet = workbook[sheet_names[2]]  # Select the third sheet
    application_detailed_spec_data = []

    # Define the start and end keywords
    end_keyword = [
        'メンバ定義\n/Member Definition',
        'メンバ名\n/Member Name',
        'アクセスレベル\n/Access Level',
    ]
    start_keywords = [
        '業務分割名\n/Business Division Name',
        '処理名\n/Process Name',
        '説明\n/Description',
        '引数\n/Argument',
        '戻り値\n/Return Value',
        'テーブル/ファイル\n/Table/File',
    ]
    start_found = False

    # Iterate through rows and collect rows between the keywords
    for row in sheet.iter_rows(values_only=True):
        filtered_row = [cell for cell in row if cell is not None]
        if any(sk in str(cell) for sk in start_keywords for cell in filtered_row):
            start_found = True
        if any(ek in str(cell) for ek in end_keyword for cell in filtered_row):
            start_found = False
        if start_found:
            if filtered_row != []:
                application_detailed_spec_data.append(filtered_row)

    return application_detailed_spec_data


def convert_spec_to_json(
    app_detailed_spec_data: list, file_path: str
) -> tuple[str, list[str]]:
    """Convert parsed app detailed spec data into JSON format.

    Returns:
        Tuple of (json_string, list_of_task_names).
    """
    is_description = False
    is_new_method = False
    is_inner_description = False
    is_process_name = False
    is_argument = False
    is_return_value = False
    is_table_file = False
    business_division_name = ""
    counter = -1
    task_names = []
    methods = [{
        "Process Name": [],
        "Argument": [],
        "Return Value": [],
        "Table or File use": [],
        "Description": []
    }]
    keywords = [
        '業務分割名\n/Business Division Name',
        '説明\n/Description',
        '処理名\n/Process Name',
        '引数\n/Argument',
        '戻り値\n/Return Value',
        'テーブル/ファイル\n/Table/File',
    ]

    # Initialize the JSON structure
    app_detailed_spec_json = {
        "Business Division Name": [],
        "Descriptions": [],
        "Methods": []
    }

    for row in app_detailed_spec_data:
        if keywords[0] in row:
            if keywords[1] not in row:
                if len(row) == 2:
                    business_division_name = row[1]
            # Extract screen/component name from file path
            base_name = extract_screen_name(file_path)
            function_name = base_name + "_" + business_division_name

            # Collect task names for the caller
            task_names.append(function_name)
            app_detailed_spec_json["Business Division Name"].append(function_name)

        if (keywords[1] in row and '名称\n/Name' not in row) or is_description:
            if keywords[2] not in row:
                is_description = True
                if len(row) == 2:
                    app_detailed_spec_json["Descriptions"].append(row[1])
                elif len(row) == 1 and row[0] != '説明\n/Description':
                    app_detailed_spec_json["Descriptions"].append(row[0])

        if keywords[2] in row or is_new_method:
            is_description = False
            is_new_method = True

            if keywords[2] in row or is_process_name:
                is_process_name = True
                is_inner_description = False
                if len(row) == 3:
                    counter += 1
                    while len(methods) < counter + 1:
                        methods.append({
                            "Process Name": [],
                            "Argument": [],
                            "Return Value": [],
                            "Table or File use": [],
                            "Description": []
                        })
                    methods[counter]["Process Name"].append(row[2])

            if keywords[3] in row or is_argument:
                # Extract the argument
                is_process_name = False
                if keywords[4] not in row:
                    is_argument = True
                    if len(row) == 4:
                        if '名称\n/Name' not in row:
                            methods[counter]["Argument"].append({
                                "No": row[0],
                                "Name": row[1],
                                "Type": row[2],
                                "Description": row[3]
                            })

            if keywords[4] in row or is_return_value:
                # Extract the return value
                is_argument = False
                if keywords[5] not in row:
                    is_return_value = True
                    if len(row) == 4:
                        if '名称\n/Name' not in row:
                            methods[counter]["Return Value"].append({
                                "No": row[0],
                                "Name": row[1],
                                "Type": row[2],
                                "Description": row[3]
                            })

            if keywords[5] in row or is_table_file:
                is_return_value = False
                if keywords[1] not in row:
                    is_table_file = True
                    if len(row) == 7:
                        methods[counter]["Table or File use"].append({
                            "No": row[0],
                            "Table_ID/File_ID": row[1],
                            "Table_Name/File_Name": row[2],
                            "CRUD Access for C": row[3],
                            "CRUD Access for R": row[4],
                            "CRUD Access for U": row[5],
                            "CRUD Access for D": row[6]
                        })

            if (keywords[1] in row and '名称\n/Name' not in row) or is_inner_description:
                is_table_file = False
                is_inner_description = True
                if len(row) == 2:
                    methods[counter]["Description"].append(row[1])

    for method in methods:
        app_detailed_spec_json["Methods"].append(method)

    # Convert to JSON string for readability
    json_string = json.dumps(app_detailed_spec_json, indent=6, ensure_ascii=False)
    print(json_string)
    return (json_string, task_names)
