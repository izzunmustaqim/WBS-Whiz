import json
import os
import shutil
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkcalendar import DateEntry
import requests  # Import pandas
import pandas as pd  # Import pandas
import config   # Import the config file
import webbrowser
import re
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from threading import Thread  # Import Thread class
import base64  # Import base64 module
from PIL import Image  # Import Image class from PIL module
import win32com.client as win32  # Import win32com.client module
import pywintypes  # Import pywintypes module

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid()
        self.skillset_file = None  # Initialize skillset_file
        self.task_details_file = None  # Initialize task_file
        self.create_widgets()
        self.skill_set_data = None  # Initialize skill_set_data
        self.task_details_data = None
        self.ss_folder_file = []
        self.screen_layout_json = None
        self.app_detailed_spec_data_converted_json = None
        self.flowchart_image_path = None
        self.task_details_response = NotImplemented
        self.task_list = []
        self.is_file_valid = True

    def create_widgets(self):
        # Add a new entry for the API key
        tk.Label(self, text="API Key").grid(row=0, column=0, padx=10, pady=5, sticky='w')
        self.api_key_entry = tk.Entry(self, width=40, show='*')
        self.api_key_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')
        
        # Members skillset input
        tk.Label(self, text="Members skill set").grid(row=2, column=0, padx=10, pady=5, sticky='w')
        self.skillset_entry = tk.Text(self, height=1, width=40)
        self.skillset_entry.grid(row=2, column=1, padx=10, pady=5, sticky='we')
        
        # btn_skillset = tk.Button(self, text="Browse", command=lambda e=self.skillset_entry, l="Members skill set": self.browse_file(e, l), width=10).grid(row=2, column=2, padx=10, pady=5)
        self.btn_skillset = tk.Button(self, text="Browse", command=lambda e=self.skillset_entry, l="Members skill set": self.browse_file(e, l), width=10)
        self.btn_skillset.grid(row=2, column=2, padx=10, pady=5)
        
        # Template download link
        text_widget = tk.Text(self, height=1, width=40, font=("Helvetica", 8), bd=0, bg=self.cget("bg"))
        text_widget.grid(row=3, column=1, padx=10, sticky='w')
        text_widget.insert(tk.END, "Download template here: Members_skillset.xlsx")
        text_widget.tag_add("link", "1.23", "1.end")
        text_widget.tag_config("link", foreground="blue", underline=True)
        text_widget.tag_bind("link", "<Button-1>", lambda e, url="https://fujitsu.sharepoint.com/:x:/r/teams/Asia-42f6e454-ChatAIContestAPG/Shared%20Documents/ChatAI%20Contest%20APG/Deliverable/Sprint%201/MEMBERS_SKILLSET.xlsx?d=w3087fb3ba54e43bab309789ad185a9a7&csf=1&web=1&e=5ekEi5": self.open_url(url))

        # Add radio buttons for Task Details and SS
        tk.Label(self, text="Input Details").grid(row=5, column=0, padx=10, pady=5, sticky='w')
         # Text area to display all selected files
        self.input_details_entry = tk.Text(self, height=5, width=40)
        self.input_details_entry.grid(row=5, column=1, padx=10, pady=5)
        
        # btn_ss_documents = tk.Button(self, text="Browse", command=lambda e=self.input_details_entry, l="SS Documents": self.browse_file(e, l), width=10).grid(row=5, column=2, padx=10, pady=5)
        self.btn_ss_documents = tk.Button(self, text="Browse", command=lambda e=self.input_details_entry, l="SS Documents": self.browse_file(e, l), width=10)
        self.btn_ss_documents.grid(row=5, column=2, padx=10, pady=5)
        
        # Template download link
        # text_widget = tk.Text(self, height=1, width=40, font=("Helvetica", 8), bd=0, bg=self.cget("bg"))
        # text_widget.grid(row=7, column=1, padx=10, sticky='w')
        # text_widget.insert(tk.END, "Download template here: Task_details.xlsx")
        # text_widget.tag_add("link", "1.23", "1.end")
        # text_widget.tag_config("link", foreground="blue", underline=True)
        # text_widget.tag_bind("link", "<Button-1>", lambda e, url="https://fujitsu.sharepoint.com/:x:/r/teams/Asia-42f6e454-ChatAIContestAPG/Shared%20Documents/ChatAI%20Contest%20APG/Deliverable/Sprint%202/Task%20Details%20Sample.xlsx?d=w4ffbd7b446c146539859793651360c36&csf=1&web=1&e=jUauGt": self.open_url(url))

        # Start and End date picker
        tk.Label(self, text="Project Duration").grid(row=9, column=0, padx=10, pady=10, sticky='w')
        #Start date
        tk.Label(self, text="Start:").grid(row=9, column=1, padx=8, pady=10, sticky='w')
        self.start_date_entry = DateEntry(self, width=12, background='darkblue', foreground='white', borderwidth=2, mindate=datetime.now(), date_pattern='MM/dd/yyyy')
        #print(self.start_date_entry.get_date()) --> to get the date value self.start_date_entry.get_date()
        self.start_date_entry.grid(row=9, column=1, padx=(50,0), pady=10, sticky='w')
        #End date
        tk.Label(self, text="End:").grid(row=9, column=1, padx=(160,0), pady=10, sticky='w')
        self.end_date_entry = DateEntry(self, width=12, background='darkblue', foreground='white', borderwidth=2, mindate=datetime.now(), date_pattern='MM/dd/yyyy')
        self.end_date_entry.grid(row=9, column=1, padx=(200,0), pady=10, sticky='w')

        # Bind the validation function to the end date entry
        self.end_date_entry.bind("<<DateEntrySelected>>", self.validate_dates)

        # Start and Cancel button
        # btn_start = tk.Button(self, text="Start", command=self.button_starter, width=10).grid(row=11, column=1, padx=10, pady=10, sticky='e')
        self.btn_start = tk.Button(self, text="Start", command=self.button_starter, width=10)
        self.btn_start.grid(row=11, column=1, padx=10, pady=10, sticky='e')
        
        tk.Button(self, text="Cancel", command=self.master.destroy, width=10).grid(row=11, column=2, padx=10, pady=10, sticky='w')
    
       

    def validate_dates(self, event):
        start_date = self.start_date_entry.get_date()
        end_date = self.end_date_entry.get_date()
        if end_date < start_date:
            messagebox.showerror("Error", "End date cannot be before start date")
            self.end_date_entry.set_date(start_date)

    def update_file_selection(self):
        self.input_details_entry.config(state=tk.NORMAL)
        self.input_details_entry.delete(1.0, tk.END)
        self.input_details_entry.config(state=tk.DISABLED)

    def create_result_section(self):
        # Add a separator
        self.separator = ttk.Separator(self, orient='horizontal')
        self.separator.grid(row=12, column=0, columnspan=3, sticky='we', pady=10)

        # Result section
        self.result_section = tk.Label(self, text="Result: ")
        self.result_section.grid(row=13, column=0, columnspan=3, padx=5, pady=5)

        # Status label to show process completion
        self.status_label = tk.Label(self, text="", wraplength=400)
        self.status_label.grid(row=14, column=0, columnspan=3, padx=10, pady=5)
        self.status_label.update_idletasks()  # Force the GUI to update

        # Add a button to save a file and center it
        self.master.grid_columnconfigure(0, weight=1)
        self.master.grid_columnconfigure(2, weight=1)
        self.download_button = tk.Button(self, text="Download WBS", command=self.download_result, width=10)
        self.download_button.grid(row=15, column=1, padx=10, pady=10, sticky='ew')
        self.download_button.grid_remove()

        self.progress = ttk.Progressbar(self, orient="horizontal", length=300, mode="determinate")
        self.progress.grid(row=16, column=1, padx=10, pady=10, sticky='ew')

    def remove_result_section(self):
        if hasattr(self, 'separator'):
            self.separator.grid_forget()
        if hasattr(self, 'result_section'):
            self.result_section.grid_forget()
        if hasattr(self, 'status_label'):
            self.status_label.grid_forget()
        if hasattr(self, 'download_button'):
            self.download_button.grid_remove()
        if hasattr(self, 'progress'):
            self.progress.grid_forget()

    def browse_file(self, entry, label): 
        # if the process is repeated - clear the list first
        self.ss_folder_file.clear()
        #selected_file_type = self.file_type.get()
        file_types = [("Excel files", "*.xlsx *.xls")]
        try:
            entry.config(state=tk.NORMAL)
            entry.delete(1.0, tk.END)
            if label == "Members skill set":
                file_path = filedialog.askopenfilename(filetypes=file_types)
                if file_path:
                    if label == "Members skill set":
                        self.skillset_file = file_path
                    else:
                        self.task_details_file = file_path
                else:
                    messagebox.showerror("Error", "No file is selected.")

                entry.insert(tk.END, file_path)
                entry.config(state=tk.DISABLED)
            else:
                # Select a folder
                entry.config(state=tk.NORMAL)
                entry.delete(1.0, tk.END)
                folder_path = filedialog.askdirectory()
                
                if folder_path:
                    self.ss_document_folder = folder_path
                    # List all file paths in the selected folder and sub-folders
                    folder_file_paths = []
                    for root, _, files in os.walk(folder_path):
                        for file in files:
                            if file.endswith('.xlsx') or file.endswith('.xls'):
                                folder_file_paths.append(os.path.join(root, file))

                    if len(folder_file_paths) > 50:
                        messagebox.showerror("Error", config.error_message["ManyExcelError"])
                        self.ss_folder_file.clear()  # Clear the list to prevent further processing
                        return
                    else:
                        print("Files in the selected folder and sub-folders:")
                        for file_path in folder_file_paths:
                            entry.insert(tk.END, file_path + "\n")
                            self.ss_folder_file.append(file_path)
                    entry.config(state=tk.DISABLED)
                else:
                    messagebox.showerror("Error", config.error_message["FolderNotFoundError"])
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while selecting the files: {e}")

    def open_url(self, url):
        webbrowser.open_new(url)

    def button_starter(self):
            t = Thread(target=self.main)
            t.start()
            # Simulate a long-running process
            
    
    def main(self):
        # if repeat the process, clear the result section first
        self.remove_result_section()
        

        self.api_key = self.validate_api_key(self.api_key_entry.get())
        if not self.api_key:
            return False
        
        # Compare excel with template
        # if self.compare_excel(self.skillset_file, 'MEMBERS_SKILLSET.xlsx') == False:
        #     return
        
        self.btn_start["state"] = tk.DISABLED
        self.btn_skillset["state"] = tk.DISABLED
        self.btn_ss_documents["state"] = tk.DISABLED
        
        # Read skillset data
        self.skill_set_data = self.read_file(self.skillset_file, 3, None)
        if self.skill_set_data is None:
            return  # Stop the process if the file is not found or an error occurred
        
        # read and return list of task
        tasks_json = self.read_ss_folder_files()

        # send first request to get complexity and priority
        self.request_task_details(tasks_json)

        # send second request to get wbs details
        self.send_data_to_chatai()
        
        self.btn_start["state"] = tk.NORMAL
        self.btn_skillset["state"] = tk.NORMAL
        self.btn_ss_documents["state"] = tk.NORMAL
    
    def compare_excel(self, file, template_file): 
        try:
            workbook1 = load_workbook(file, data_only=True)
            workbook2 = load_workbook(template_file, data_only=True)
            sheet1 = workbook1.active
            sheet2 = workbook2.active

            comparison1_success = True
            comparison2_success = True
            comparison3_success = True
    
            if workbook1 == self.task_details_file:
                # Comparison 1: Column B, rows 2-4
                for row in range(2, 5):  # Rows 2, 3, 4
                    if sheet1[f"B{row}"].value != sheet2[f"B{row}"].value:
                        print(f"Comparison 1 failed: Cell B{row} differs.")
                        comparison1_success = False
    
                # Comparison 2: Columns B-E, row 6
            for col in range(2, 6):  # Columns B, C, D, E (2,3,4,5)
                if sheet1[f"{chr(ord('B') + col - 2)}{6}"].value != sheet2[f"{chr(ord('B') + col - 2)}{6}"].value:
                    print(f"Comparison 2 failed: Cell {chr(ord('B') + col - 2)}{6} differs.")
                    comparison2_success = False
    
            if workbook1 == self.skillset_file:
                # Comparison 3: Columns B-GP, rows 2-4
                for row in range(2, 5):  # Rows 2, 3, 4
                    for col in range(2, 188):  # Columns B (2) to GP (187)
                        col_str = chr(col + 64)  # Convert column number to letter
                        if sheet1.cell(row=row, column=col).value != sheet2.cell(row=row, column=col).value:
                            print(f"Comparison 3 failed: Cell {col_str}{row} differs.")
                            comparison3_success = False

            overall_comparison = comparison1_success and comparison2_success and comparison3_success

            print(f"Overall comparison: {overall_comparison}")
            if overall_comparison != True:
                messagebox.showerror("Overall Comparison", f"Please double check format/input in excel")
                return overall_comparison
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to compare Excel files: {e}")
            return False
        
    # rangecol=None mean will read all columns that has data
    def read_file(self, file, rowskip=0, rangecol=None):
        # print(file, rowskip, rangecol)
        try:
            file_size = os.path.getsize(file)
            max_mb = 25 * 1024 * 1024  # 25MB in bytes
            if file_size > max_mb:
                messagebox.showerror("Error","The member skillset file is too large. Please upload a file smaller than 25MB and try again")
                self.browse_file(self.skillset_entry, "Members skill set")
                self.btn_start["state"] = tk.NORMAL
                self.btn_skillset["state"] = tk.NORMAL
                self.btn_ss_documents["state"] = tk.NORMAL
                sys.exit()


            file_data = pd.read_excel(file, skiprows=rowskip, usecols=rangecol)  # Read Excel file using pandas
            
            # Check if the file is empty
            if file_data.empty:
                raise pd.errors.EmptyDataError("The file is empty")
            
            # Data cleaning steps
            file_data.dropna(axis=1, how='all', inplace=True)  # Remove columns with all missing values
            file_data.fillna(0, inplace=True)  # Replace NaN values with 0
            
            # print(file_data)  # Print the task details data for debugging
            return(file_data)
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except pd.errors.EmptyDataError:
            messagebox.showerror("Error", config.error_message["EmptyDataError"])
            return None
        except pd.errors.ParserError:
            messagebox.showerror("Error", config.error_message["ParserError"])
            return None
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
            return None
    
    def process_step(self):
        current_value = self.progress["value"]
        if current_value < 100:
            self.progress["value"] = current_value + 1
        else:
            self.progress["value"] = 0  # Reset to 0 once it reaches 100
        self.master.after(100, self.process_step)  # Update every 100 milliseconds

    def read_ss_folder_files(self):
        self.is_file_valid = True
        self.selected_folder_file = []
        self.task_list.clear()
        is_ss_doc = False
        
        # Call the method to create the status section
        self.create_result_section()

        self.progress["value"] = 0
        self.status_label.config(text="Processing input data...")

        # # Simulate sending data to ChatAI
        self.master.after(100, self.process_step)

        # Variables to store file paths based on keywords
        application_detailed_specification_files = []
        event_process_diagram_history_files = []
        screen_layout_files = []
        
        
        for file_path in self.ss_folder_file:

            if "Application Detailed Specification" in file_path:
                is_ss_doc = True
                application_detailed_specification_files.append(file_path)
                
            # elif "Event Process Sequence Diagram History" in file_path:
            #     event_process_diagram_history_files.append(file_path)

            elif "Screen Layout" in file_path:
                is_ss_doc = True
                screen_layout_files.append(file_path)
                # Define the regular expression pattern to match the desired part   

        if is_ss_doc == False:
            messagebox.showerror("Error", "No SS documents found in the folder. Please choose the correct folder and try again")
            self.browse_file(self.input_details_entry, "SS Documents")
            self.btn_start["state"] = tk.NORMAL
            self.btn_skillset["state"] = tk.NORMAL
            self.btn_ss_documents["state"] = tk.NORMAL
            self.remove_result_section()
            sys.exit()
                
        # Print the results
        print("\nScreen Layout Files:")
        for file in screen_layout_files:

            sheetName = "項目定義"
            keywordsHeader = ['画面項目名\n/Screen Item Name', 'タイプ\n/ Type']

            # to check file validity
            workbook = load_workbook(filename=file)
            sheet_names = [] # Initialize an empty list for sheet names

            # Iterate through each worksheet in the workbook
            for sheet in workbook.worksheets:
                sheet_names.append(sheet.title)

            # Check if the file is an screen layout file
            error_message = self.check_file_validity(sheet_names, workbook, file)

            if self.is_file_valid:
                
                self.screen_layout_json = self.read_screen_layout(file, sheetName, keywordsHeader)
                print(f"{self.screen_layout_json}")
            
            else:
                messagebox.showerror("Error", f"{error_message}")
                self.browse_file(self.input_details_entry, "SS Documents")
                self.btn_start["state"] = tk.NORMAL
                self.btn_skillset["state"] = tk.NORMAL
                self.btn_ss_documents["state"] = tk.NORMAL
                self.remove_result_section()
                sys.exit()

        print("Application Detailed Specification Files:")
        for file in application_detailed_specification_files:
            # print(file)
            # call function read application detailed specification files
            workbook = load_workbook(filename=file_path, data_only=True) # Load the workbook
            sheet_names = [] # Initialize an empty list for sheet names

            # Iterate through each worksheet in the workbook
            for sheet in workbook.worksheets:
                sheet_names.append(sheet.title)

            # Check if the file is an Application Detailed Specification file
            error_message = self.check_file_validity(sheet_names, workbook, file_path)

            if self.is_file_valid:
                app_detailed_spec_data = self.read_application_detailed_specification_files(file)
                self.app_detailed_spec_data_converted_json = self.convert_app_detailed_spec_data(app_detailed_spec_data, file)
                #print(self.app_detailed_spec_data_converted_json)
            
            else:
                messagebox.showerror("Error", f"{error_message}")
                self.browse_file(self.input_details_entry, "SS Documents")
                self.btn_start["state"] = tk.NORMAL
                self.btn_skillset["state"] = tk.NORMAL
                self.btn_ss_documents["state"] = tk.NORMAL
                self.remove_result_section()
                sys.exit()
                

        # create a list of tasks into json
        json_list = []
        json_list = [{'Item No': i + 1, 'Task Description': task} for i, task in enumerate(self.task_list)]
        # Convert to JSON string
        json_string = json.dumps(json_list, indent=4)
        self.progress["value"] = 100

        print(json_string)
        return json_string

    def read_screen_layout(self, file, sheetName, keywordsHeader):
        try:
            workbook = load_workbook(filename=file)
            sheet_names = [] # Initialize an empty list for sheet names

            # Iterate through each worksheet in the workbook
            for sheet in workbook.worksheets:
                sheet_names.append(sheet.title)

            # Check if the file is an Application Detailed Specification file
            # self.check_file_validity(sheet_names, workbook, file)

            sheet = workbook[sheetName]  
            screen_layout_data = []
            start_found = False

            for row in sheet.iter_rows(values_only=True):
                filtered_row = []
                for cell in row:
                    if cell is not None:
                        filtered_row.append(cell)

                if any(keyword in str(cell) for keyword in keywordsHeader for cell in filtered_row):
                    start_found = True
                    
                if start_found:
                    if filtered_row != [] and '画面項目名\n/Screen Item Name' not in filtered_row:
                        screen_layout_data.append(filtered_row)
            
            screen_name = ""
            match = re.search(r"\\([^_]+)_", file)
            if match:
                    extracted_part = match.group(1)
            
            # Check if there is '\\' inside the string and count occurrences
            backslash_count = extracted_part.count('\\')

            # Loop to extract the part after the last backslash if there are multiple backslashes
            if backslash_count > 0:
                parts = extracted_part.split('\\')
                screen_name = parts[-1] + "_UI"
            else:
                screen_name = extracted_part + "_UI"

            # creating the list of tasks
            self.task_list.append(screen_name)

            # Initialize the JSON structure
            screen_layout_json = {
                screen_name: [],
            }
    
            for row in screen_layout_data:
                if len(row) > 1 and (row[1] != '-' or row[2] != '-'):
                    screen_layout_json[screen_name].append({
                        "Screen Item Name": row[1],
                        "Type": row[2],
                    })
            
            # Convert to JSON string for readability
            json_string = json.dumps(screen_layout_json, indent=4)
            return json_string
    
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except ValueError as e:
            messagebox.showerror("Error", config.error_message["EmptyDataError"])
            return None
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            return None

    def read_application_detailed_specification_files(self, file_path):
        # Read application detailed specification files
        try:
            workbook = load_workbook(filename=file_path, data_only=True) # Load the workbook
            sheet_names = [] # Initialize an empty list for sheet names

            # Iterate through each worksheet in the workbook
            for sheet in workbook.worksheets:
                sheet_names.append(sheet.title)

            # Check if the file is an Application Detailed Specification file
            # self.check_file_validity(sheet_names, workbook, file_path)
            
            sheet = workbook[sheet_names[2]] # Select the third sheet
            application_detailed_spec_data = []

            # Define the start and end keywords
            end_keyword = ['メンバ定義\n/Member Definition','メンバ名\n/Member Name', 'アクセスレベル\n/Access Level']
            start_keywords = ['業務分割名\n/Business Division Name','処理名\n/Process Name','説明\n/Description', '引数\n/Argument', '戻り値\n/Return Value', 'テーブル/ファイル\n/Table/File']
            start_found = False     # Initialize flags

            # Iterate through rows and print rows between the keywords
            for row in sheet.iter_rows(values_only=True):
                filtered_row = [cell for cell in row if cell is not None]
                # print(filtered_row)
                if any(start_keyword in str(cell) for start_keyword in start_keywords for cell in filtered_row):
                    start_found = True
                if any(end_keyword in str(cell) for end_keyword in end_keyword for cell in filtered_row):
                    start_found = False
                if start_found:
                    if filtered_row != []:
                        #print(filtered_row)
                        application_detailed_spec_data.append(filtered_row)
            #print(application_detailed_spec_data)
            return application_detailed_spec_data
        
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except pd.errors.EmptyDataError:
            messagebox.showerror("Error", config.error_message["EmptyDataError"])
            return None
        except pd.errors.ParserError:
            messagebox.showerror("Error", config.error_message["ParserError"])
            return None
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
            return None
    
    # def event_Process_Sequence_Diagram_History_Files(self, file_path):
    #     try:
    #         sheet_names = [] # Initialize an empty list for sheet names
    #         workbook = load_workbook(file_path, data_only=True)

    #         # Iterate through each worksheet in the workbook
    #         for sheet in workbook.worksheets:
    #             sheet_names.append(sheet.title)
    #             print(sheet.title)

    #         # # Check if the file is an Application Detailed Specification file
    #         self.check_file_validity(sheet_names, workbook, file_path)
            
    #         sheet = workbook[sheet_names[2]] # Select the third sheet

    #         # Extract images from the Excel file
    #         images = []
    #         for idx, image in enumerate(sheet._images):
    #             # Convert the openpyxl image to a PIL image
    #             image_stream = io.BytesIO(image._data())
    #             img = Image.open(image_stream)
    #             images.append(img)
                
    #             # Save the image to a file
    #             img.save(f'image_{idx + 1}.png')

    #         image_path = "image_1.png"
    #         #description = self.describe_image(image_path)
    #         # Delete the file
    #         # os.remove(image_path)
    #         return image_path
    #         # print(description)

    #     except FileNotFoundError:
    #         messagebox.showerror("Error", config.error_message["FileNotFoundError"])
    #         return None
    #     except pd.errors.EmptyDataError:
    #         messagebox.showerror("Error", config.error_message["EmptyDataError"])
    #         return None
    #     except pd.errors.ParserError:
    #         messagebox.showerror("Error", config.error_message["ParserError"])
    #         return None
    #     except Exception as e:
    #         messagebox.showerror("Error", f"Failed to read Excel file: {e}")
    #         return None

    # def describe_image(self, img):
    #     encoded_image = self.encode_image(img)
    
    #     headers = {
    #         "Content-type": "application/json",
    #         "api-key": self.api_key
    #     }
    
    #     # data = {
    #     #     "messages": [
    #     #         {"role": "user", "content": [
    #     #             {"type": "text", "text": "Please describe the image below."},
    #     #             {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{encoded_image}"}}
    #     #         ]}
    #     #     ]
    #     # }

    #     data = {
    #         "contents": [
    #             {
    #                 "role": "user",
    #                 "parts": [
    #                     {
    #                         "inlineData": {
    #                             "mimeType": "image/jpeg",
    #                             "data": encoded_image
    #                         }
    #                     },
    #                     {
    #                         "text": "What is shown this image?"
    #                     }
    #                 ]
    #             }
    #         ]
    #     }
    
    #     # response = requests.post("https://ai-foundation-api.app/ai-foundation/chat-ai/gpt4", headers=headers, json=data)
    #     response = requests.post("https://api.ai-service.global.fujitsu.com/ai-foundation/chat-ai/gemini/pro:generateContent", headers=headers, json=data)
    #     response_json = response.json()
    #     # print(response_json)
    #     # return response_json["choices"][0]["message"]["content"]
    #     return response_json['candidates'][0]['content']['parts'][0]['text']

    # def encode_image(self,img):
    #     with open(img, "rb") as image_file:
    #         return base64.b64encode(image_file.read()).decode('utf-8')
    
    def check_file_validity(self, sheet_names, workbook, file_path):
        error_msg = ""
        file_size = os.path.getsize(file_path)
        max_mb = 25 * 1024 * 1024  # 25MB in bytes
        if file_size > max_mb:
            # raise ValueError("The file is too large. Please upload a smaller file size")
            error_msg = "Error! The SS file is too large. Please upload a file smaller than 25MB"
            self.is_file_valid=False
            return error_msg
            # print(f"File size: {file_size} bytes")
        for sheet_name in sheet_names:
                sheet_temp = workbook[sheet_name]
                if sheet_temp.max_row == 1 and sheet_temp.max_column == 1 and sheet_temp.cell(row=1, column=1).value is None:
                    # raise ValueError("The file is empty")
                    error_msg = "Error! The screen layout files are empty. Please upload the correct file"
                    # self.browse_file(self.input_details_entry, "SS Documents")
                    self.is_file_valid=False
                    return error_msg
                for row in sheet_temp.iter_rows(values_only=True):
                    filtered_row = [cell for cell in row if cell is not None]
                    if filtered_row == []:
                        error_msg = "Error! The screen layout files are empty. Please upload the correct file and start the process again"
                        # self.browse_file(self.input_details_entry, "SS Documents")
                        self.is_file_valid=False
                        return error_msg

                    elif "Screen Layout" in file_path and '画面レイアウト\n/Screen Layout' not in filtered_row:
                        # raise ValueError("The file is not a Screen Layout file")
                        error_msg = "Error! The file is not a Screen Layout file. Please upload the correct file and start the process again"
                        # self.browse_file(self.input_details_entry, "SS Documents")
                        self.is_file_valid=False
                        return error_msg

                    elif "Application Detailed Specification" in file_path and "アプリケーション詳細仕様\n/Application Detailed Specification" not in filtered_row:
                        # raise ValueError("The file is not an Application Detailed Specification file")
                        error_msg = "Error! The file is not an Application Detailed Specification file. Please upload the correct file and start the process again"
                        # self.browse_file(self.input_details_entry, "SS Documents")
                        self.is_file_valid=False
                        return error_msg

                    elif "Event Process Sequence Diagram History" in file_path and "イベント処理シーケンス図\n/Event Process Sequence Diagram" not in filtered_row:
                        # raise ValueError("The file is not an Event Process Sequence Diagram History file")
                        error_msg == "Error! The file is not an Event Process Sequence Diagram History file. Please upload the correct file and start the process again"
                        # self.browse_file(self.input_details_entry, "SS Documents")
                        self.is_file_valid=False
                        return error_msg

                    else:
                        return "File is valid"
        
    def convert_app_detailed_spec_data(self, app_detailed_spec_data, file_path):
        is_description = False
        is_new_method = False
        is_inner_description = False
        is_process_name = False
        is_argument = False
        is_return_value = False
        is_table_file = False
        business_division_name = ""
        counter = -1
        methods = [{
            "Process Name": [], 
            "Argument": [], 
            "Return Value": [], 
            "Table or File use": [],
            "Description": []
            }]
        keywords = ['業務分割名\n/Business Division Name', '説明\n/Description', '処理名\n/Process Name', '引数\n/Argument', '戻り値\n/Return Value', 'テーブル/ファイル\n/Table/File']

        # Initialize the JSON structure
        app_detailed_spec_json = {
            "Business Division Name": [],
            "Descriptions": [],
            "Methods": []
        }

        for row in app_detailed_spec_data:
            #print(row)
            if keywords[0] in row:
                #is_table_file = False
                if keywords[1] not in row:
                    if len(row) == 2:
                        business_division_name = row[1]
                # Define the regular expression pattern to match the desired part
                function_name = ""
                # Use re.search to find the match
                #print(file_path)
                match = re.search(r"\\([^_]+)_", file_path)
                if match:
                        extracted_part = match.group(1)
                
                # Check if there is '\\' inside the string and count occurrences
                backslash_count = extracted_part.count('\\')

                # Loop to extract the part after the last backslash if there are multiple backslashes
                if backslash_count > 0:
                    parts = extracted_part.split('\\')
                    function_name = parts[-1] + "_" + business_division_name
                else:
                    function_name = extracted_part + "_" + business_division_name
                
                # creating the list of tasks
                self.task_list.append(function_name)
                app_detailed_spec_json["Business Division Name"].append(function_name)
                

            if (keywords[1] in row and '名称\n/Name' not in row) or is_description:
                if keywords[2] not in row:
                    is_description = True
                    if len(row) == 2:
                        #description.append(row[1])
                        app_detailed_spec_json["Descriptions"].append(row[1])
                    elif len(row) == 1 and row[0] != '説明\n/Description':
                        #description.append(row[0])
                        app_detailed_spec_json["Descriptions"].append(row[0])
                

            if keywords[2] in row or is_new_method:
                is_description = False
                is_new_method = True

                if keywords[2] in row or is_process_name:
                    is_process_name = True
                    is_inner_description = False
                    if len(row) == 3:
                        counter += 1
                        while len(methods) < counter+1:
                            methods.append({
                                "Process Name": [],
                                "Argument": [],
                                "Return Value": [],
                                "Table or File use": [],
                                "Description": []
                            })
                        methods[counter]["Process Name"].append(row[2])
                
                if keywords[3] in row or is_argument:
                    # Extract the argument
                    is_process_name = False
                    if keywords[4] not in row:
                        is_argument = True
                        if len(row) == 4:
                            if '名称\n/Name' not in row:
                                #argument.append(row)
                                methods[counter]["Argument"].append({
                                                                            "No": row[0],
                                                                            "Name": row[1],
                                                                            "Type": row[2],
                                                                            "Description": row[3]
                                                                        })

                if keywords[4] in row or is_return_value:
                    # Extract the return value
                    is_argument = False
                    if keywords[5] not in row:
                        is_return_value = True
                        if len(row) == 4:
                            if '名称\n/Name' not in row:
                                #return_value.append(row)
                                methods[counter]["Return Value"].append({
                                                                                "No": row[0],
                                                                                "Name": row[1],
                                                                                "Type": row[2],
                                                                                "Description": row[3]
                                                                            })

                if keywords[5] in row or is_table_file:
                    #print(row)
                    is_return_value = False
                    if keywords[1] not in row:
                        is_table_file = True
                        if len(row) == 7:
                            methods[counter]["Table or File use"].append({
                                                                                "No": row[0],
                                                                                "Table_ID/File_ID": row[1],
                                                                                "Table_Name/File_Name": row[2],
                                                                                "CRUD Access for C": row[3],
                                                                                "CRUD Access for R": row[4],
                                                                                "CRUD Access for U": row[5],
                                                                                "CRUD Access for D": row[6]
                                                                            })
                
                if (keywords[1] in row and '名称\n/Name' not in row) or is_inner_description:
                    #print(row)
                    is_table_file = False
                    is_inner_description = True
                    if len(row) == 2:
                        methods[counter]["Description"].append(row[1])
                
        for method in methods:
            app_detailed_spec_json["Methods"].append(method)
                
        # Convert to JSON string for readability
        json_string = json.dumps(app_detailed_spec_json, indent=6, ensure_ascii=False)
        print(json_string)
        return json_string

    
    
    def request_task_details(self, tasks_list_json):
        try:
            # # Call the method to create the status section
            # self.create_result_section()

            self.progress["value"] = 0
            self.status_label.config(text="Sending request to get task complexity...")

            # # Simulate sending data to ChatAI
            self.master.after(100, self.process_step)

            # encoded_image = self.encode_image(self.flowchart_image_path)
    
            headers = {
                "Content-type": "application/json",
                "api-key": self.api_key
            }
            prompt = config.prompt_list_task.format(
                            screen_layout_json=self.screen_layout_json,
                            app_detailed_spec_data_converted_json=self.app_detailed_spec_data_converted_json,
                            tasks_list_json=tasks_list_json
                        )
            api_endpoint = "https://api.ai-service.global.fujitsu.com/ai-foundation/chat-ai/gemini/pro:generateContent" 

            data = {
                "contents": [
                    {
                        "role": "user",
                        "parts": [
                            {
                                "text": prompt
                            }
                        ]
                    }
                ]
            }
        
            # Send the POST request
            response = requests.post(api_endpoint, headers=headers, json=data)
            response.raise_for_status()  # Raise an exception for HTTP errors
            # os.remove(self.flowchart_image_path)
           
            # Check the response
            try:
                analysis_result = response.json()
                # print("Analysis Result:", analysis_result)
                # Extract the content (only the wbs result)
                content = analysis_result['candidates'][0]['content']['parts'][0]['text']
                self.task_details_response = content
                print("Response from chat AI for the Tasks Complexity")
                print(content)
                
                # Extract the markdown table using regular expression
                # table_pattern = re.compile(r'\|.*\|')
                # markdown_table = '\n'.join(table_pattern.findall(content))
                
                # # Convert the Markdown table to a DataFrame
                # data = io.StringIO(markdown_table)
                # task_df = pd.read_csv(data, sep="|", skipinitialspace=True, engine='python')

                # Remove the first row
                # self.task_df = task_df.iloc[1:]

                # Drop the first and last columns which are empty due to the table format
                # self.task_df = task_df.drop(task_df.columns[[0, -1]], axis=1)                
 
            except json.JSONDecodeError:
                print("Error: The response is not in JSON format.")
                print("Response content:", response.text)
 
        except requests.exceptions.RequestException as e:
            if "Too Large" in str(e):
                messagebox.showerror("Error", config.error_message["FileTooBig"])
            else:
                print("Failed to get a response from ChatAI. Status code:", response.status_code)
                print("Response content:", response.text)
                messagebox.showerror("Error", f"Failed to send data to ChatAI: {e}")

        except ValueError as ve:
            print(ve)
            messagebox.showerror("Error", str(ve))
        finally:
            # self.status_label.config(text="Process Task Details has completed successfully. Creating WBS is in progress.")
            self.progress["value"] = 100
            

    # Send data for wbs - request 2
    def send_data_to_chatai(self):
        start_date_str=self.start_date_entry.get_date()
        end_date_str=self.end_date_entry.get_date()
        start_date = start_date_str.strftime('%m/%d/%Y')
        end_date = end_date_str.strftime('%m/%d/%Y')
        
        try:   
            task_details_data=self.task_details_response
            # for progress bar
            self.progress["value"] = 0
            self.status_label.config(text="Process Task Details has completed successfully. Sending request to get the WBS details...")

            # Simulate sending data to ChatAI
            self.master.after(100, self.process_step)

            # Define the API endpoint and hardcoded prompt
            api_endpoint = "https://api.ai-service.global.fujitsu.com/ai-foundation/chat-ai/gemini/pro:generateContent" 
            prompt = config.prompt.format(
                            task_details_data=task_details_data,
                            # task_details_data=self.task_df.to_json(),
                            skill_set_data=self.skill_set_data.to_json(),
                            start_date_str=start_date,
                            end_date_str=end_date,
                            task_description="Task Description Example",  # Provide example values for placeholders
                            assigned_to="Assigned to Example",
                            progress="To do",
                            plan_start_date="Start date Example",
                            plan_end_date="End date Example",
                        )

            headers = {
                "Content-type": "application/json",
                "api-key": self.api_key
            }

            payload = {
                "contents": [
                {
                    "role": "user",
                    "parts": [
                        {
                            "text": prompt
                        }
                    ]
                }
                ]
            }

            # Send the POST request
            response = requests.post(api_endpoint, headers=headers, json=payload)
            response.raise_for_status()  # Raise an exception for HTTP errors

            # Check the response
            try:
                analysis_result = response.json()
                #print("Analysis Result:", analysis_result)

                # Extract the content (only the wbs result)
                content = analysis_result['candidates'][0]['content']['parts'][0]['text']
                self.create_wbs(content, self.start_date_entry, self.end_date_entry)
                print(content)

            except json.JSONDecodeError:
                print("Error: The response is not in JSON format.")
                print("Response content:", response.text)

        except requests.exceptions.RequestException as e:
            if "Too Large" in str(e):
                messagebox.showerror("Error", config.error_message["FileTooBig"])
            else:
                print("Failed to get a response from ChatAI. Status code:", response.status_code)
                print("Response content:", response.text)
                messagebox.showerror("Error", f"Failed to send data to ChatAI: {e}")

        except ValueError as ve:
            print(ve)
            messagebox.showerror("Error", str(ve))
        finally:
            self.status_label.config(text="Process has completed successfully. You may download the WBS file using the download button below.")
            self.progress["value"] = 100
            self.download_button.grid()
            if hasattr(self, 'progress'):
                self.progress.grid_forget()

    def download_result(self):
        try:
            # Define the destination file path in the Downloads folder
            downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
            destination_file_path = os.path.join(downloads_folder, "Details_WBS.xlsm")

            # Create dummy file in the download folder
            df = pd.DataFrame()
            df.to_excel(destination_file_path, index=False)

            # Get the current directory
            current_directory = os.getcwd()

            # Define the source file path
            source_file_path = os.path.join(current_directory, "Details_WBS.xlsm")

            # Copy the file
            shutil.copy(source_file_path, destination_file_path)
            messagebox.showinfo("Success", f"File saved successfully to {destination_file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")

    def validate_api_key(self, api_key):
        pattern = r'^[A-Za-z0-9]{48}$'        
        if not api_key:
            messagebox.showerror("Error", config.error_message["APIEmptyField"])
            return None
        elif re.search(r'[\uFF01-\uFF60\uFFE0-\uFFE6]', api_key):
            messagebox.showerror("Error", config.error_message["FullWidthCharacterError"])
            return None
        elif re.match(pattern, api_key):
            return api_key
        else:
            messagebox.showerror("Error", config.error_message["InvalidKeyError"])
            return None
        
    def create_wbs(self, content, start_date, end_date):
        # Extract the date value
        start_date_value = start_date.get_date()
        end_date_value = end_date.get_date()

        # Extract the markdown table using regular expression
        table_pattern = re.compile(r'\|.*\|')
        markdown_table = '\n'.join(table_pattern.findall(content))

        # Convert the Markdown table to a DataFrame
        data = io.StringIO(markdown_table)
        df = pd.read_csv(data, sep="|", skipinitialspace=True, engine='python')

        # Remove the first row
        df = df.iloc[1:]

        # Drop the first and last columns which are empty due to the table format
        df = df.drop(df.columns[[0, -1]], axis=1)

        # remove the header from an existing Pandas DataFrame
        df = df.rename(columns=df.iloc[0]).drop(df.index[0])

        try:
            # Load the Excel template
            template_path = 'JDU-WBS_Template_Samples.xlsm'
            # Create an instance of Excel
            macro_name = 'UpdateDatesAndFormat'  # Name of the macro to run
            excel = win32.gencache.EnsureDispatch("Excel.Application")
    
            # Open the workbook
            workbook = excel.Workbooks.Open(os.path.abspath(template_path))
            sheet = workbook.Sheets(1)

            # Write the variable into cell
            sheet.Cells(2, 2).Value = "Details_WBS.xlsm"
            # Format the date as a string
            sheet.Cells(6, 2).Value = start_date_value.strftime('%m/%d/%Y')
            sheet.Cells(7, 2).Value = end_date_value.strftime('%m/%d/%Y')
            # Set current_date to the current date
            current_date = datetime.now().date()
            pywintypes_time = pywintypes.Time(current_date)
            sheet.Cells(2, 7).Value = pywintypes_time

            # Write the DataFrame to the Excel template starting at row 10
            start_row = 10 #temp test for no 10
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    sheet.Cells(r_idx, c_idx).Value = value
    
            # Run the macro
            excel.Application.Run(macro_name)
    
            # Save and close the workbook
            workbook.SaveAs(os.path.abspath("Details_WBS.xlsm"))
            workbook.Close()
        
            # Quit Excel
            excel.Application.Quit()

            output_path = 'Details_WBS.xlsm'
            
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except PermissionError as e:
            messagebox.showerror("Error", f"Error: Permission denied. {e}")
            return None
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            return None

        print(f"DataFrame saved to {output_path}")


root = tk.Tk()
root.title("WBS Enhancement")
app = Application(master=root)
app.mainloop()