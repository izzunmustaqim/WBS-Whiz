"""WBS Enhancement Application GUI.

The Application class handles all Tkinter GUI interactions and delegates
business logic to the file_parser, api_client, and wbs_writer modules.
"""

# Standard library
import json
import os
import re
import tkinter as tk
import webbrowser
from datetime import datetime
from threading import Thread
from tkinter import filedialog, messagebox, ttk

# Third-party
import pandas as pd
import requests
from openpyxl import load_workbook
from tkcalendar import DateEntry

# Local
import config
from file_parser import (
    read_excel_file,
    check_file_validity,
    parse_screen_layout,
    parse_app_detailed_spec,
    convert_spec_to_json,
)
from api_client import send_gemini_request
from wbs_writer import (
    markdown_table_to_dataframe,
    write_wbs_to_excel,
    copy_to_downloads,
)


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid()
        self.skillset_file = None
        self.task_details_file = None
        self.create_widgets()
        self.skill_set_data = None
        self.task_details_data = None
        self.ss_folder_file = []
        self.screen_layout_json = None
        self.app_detailed_spec_data_converted_json = None
        self.flowchart_image_path = None
        self.task_details_response = None
        self.task_list = []
        self.is_file_valid = True

    # ------------------------------------------------------------------ #
    #  Widget creation & layout                                           #
    # ------------------------------------------------------------------ #

    def create_widgets(self):
        # API key entry
        tk.Label(self, text="API Key").grid(row=0, column=0, padx=10, pady=5, sticky='w')
        self.api_key_entry = tk.Entry(self, width=40, show='*')
        self.api_key_entry.grid(row=0, column=1, padx=10, pady=5, sticky='we')

        # Members skillset input
        tk.Label(self, text="Members skill set").grid(row=2, column=0, padx=10, pady=5, sticky='w')
        self.skillset_entry = tk.Text(self, height=1, width=40)
        self.skillset_entry.grid(row=2, column=1, padx=10, pady=5, sticky='we')

        self.btn_skillset = tk.Button(
            self, text="Browse",
            command=lambda e=self.skillset_entry, l="Members skill set": self.browse_file(e, l),
            width=10,
        )
        self.btn_skillset.grid(row=2, column=2, padx=10, pady=5)

        # Template download link
        text_widget = tk.Text(self, height=1, width=40, font=("Helvetica", 8), bd=0, bg=self.cget("bg"))
        text_widget.grid(row=3, column=1, padx=10, sticky='w')
        text_widget.insert(tk.END, "Download template here: Members_skillset.xlsx")
        text_widget.tag_add("link", "1.23", "1.end")
        text_widget.tag_config("link", foreground="blue", underline=True)
        text_widget.tag_bind("link", "<Button-1>", lambda e, url="https://fujitsu.sharepoint.com/:x:/r/teams/Asia-42f6e454-ChatAIContestAPG/Shared%20Documents/ChatAI%20Contest%20APG/Deliverable/Sprint%201/MEMBERS_SKILLSET.xlsx?d=w3087fb3ba54e43bab309789ad185a9a7&csf=1&web=1&e=5ekEi5": self.open_url(url))

        # Input details (SS documents)
        tk.Label(self, text="Input Details").grid(row=5, column=0, padx=10, pady=5, sticky='w')
        self.input_details_entry = tk.Text(self, height=5, width=40)
        self.input_details_entry.grid(row=5, column=1, padx=10, pady=5)

        self.btn_ss_documents = tk.Button(
            self, text="Browse",
            command=lambda e=self.input_details_entry, l="SS Documents": self.browse_file(e, l),
            width=10,
        )
        self.btn_ss_documents.grid(row=5, column=2, padx=10, pady=5)

        # Start and End date picker
        tk.Label(self, text="Project Duration").grid(row=9, column=0, padx=10, pady=10, sticky='w')
        tk.Label(self, text="Start:").grid(row=9, column=1, padx=8, pady=10, sticky='w')
        self.start_date_entry = DateEntry(
            self, width=12, background='darkblue', foreground='white',
            borderwidth=2, mindate=datetime.now(), date_pattern='MM/dd/yyyy',
        )
        self.start_date_entry.grid(row=9, column=1, padx=(50, 0), pady=10, sticky='w')

        tk.Label(self, text="End:").grid(row=9, column=1, padx=(160, 0), pady=10, sticky='w')
        self.end_date_entry = DateEntry(
            self, width=12, background='darkblue', foreground='white',
            borderwidth=2, mindate=datetime.now(), date_pattern='MM/dd/yyyy',
        )
        self.end_date_entry.grid(row=9, column=1, padx=(200, 0), pady=10, sticky='w')

        # Bind the validation function to the end date entry
        self.end_date_entry.bind("<<DateEntrySelected>>", self.validate_dates)

        # Start and Cancel buttons
        self.btn_start = tk.Button(self, text="Start", command=self.button_starter, width=10)
        self.btn_start.grid(row=11, column=1, padx=10, pady=10, sticky='e')

        tk.Button(self, text="Cancel", command=self.master.destroy, width=10).grid(
            row=11, column=2, padx=10, pady=10, sticky='w'
        )

    # ------------------------------------------------------------------ #
    #  Result section management                                          #
    # ------------------------------------------------------------------ #

    def create_result_section(self):
        self.separator = ttk.Separator(self, orient='horizontal')
        self.separator.grid(row=12, column=0, columnspan=3, sticky='we', pady=10)

        self.result_section = tk.Label(self, text="Result: ")
        self.result_section.grid(row=13, column=0, columnspan=3, padx=5, pady=5)

        self.status_label = tk.Label(self, text="", wraplength=400)
        self.status_label.grid(row=14, column=0, columnspan=3, padx=10, pady=5)
        self.status_label.update_idletasks()

        self.master.grid_columnconfigure(0, weight=1)
        self.master.grid_columnconfigure(2, weight=1)
        self.download_button = tk.Button(self, text="Download WBS", command=self.download_result, width=10)
        self.download_button.grid(row=15, column=1, padx=10, pady=10, sticky='ew')
        self.download_button.grid_remove()

        self.progress = ttk.Progressbar(self, orient="horizontal", length=300, mode="determinate")
        self.progress.grid(row=16, column=1, padx=10, pady=10, sticky='ew')

    def remove_result_section(self):
        if hasattr(self, 'separator'):
            self.separator.grid_forget()
        if hasattr(self, 'result_section'):
            self.result_section.grid_forget()
        if hasattr(self, 'status_label'):
            self.status_label.grid_forget()
        if hasattr(self, 'download_button'):
            self.download_button.grid_remove()
        if hasattr(self, 'progress'):
            self.progress.grid_forget()

    def process_step(self) -> None:
        current_value = self.progress["value"]
        if current_value < 100:
            self.progress["value"] = current_value + 1
        else:
            self.progress["value"] = 0
        self.master.after(100, self.process_step)

    # ------------------------------------------------------------------ #
    #  User interaction handlers                                          #
    # ------------------------------------------------------------------ #

    def validate_dates(self, event):
        start_date = self.start_date_entry.get_date()
        end_date = self.end_date_entry.get_date()
        if end_date < start_date:
            messagebox.showerror("Error", "End date cannot be before start date")
            self.end_date_entry.set_date(start_date)

    def update_file_selection(self):
        self.input_details_entry.config(state=tk.NORMAL)
        self.input_details_entry.delete(1.0, tk.END)
        self.input_details_entry.config(state=tk.DISABLED)

    def browse_file(self, entry, label):
        # if the process is repeated - clear the list first
        self.ss_folder_file.clear()
        file_types = [("Excel files", "*.xlsx *.xls")]
        try:
            entry.config(state=tk.NORMAL)
            entry.delete(1.0, tk.END)
            if label == "Members skill set":
                file_path = filedialog.askopenfilename(filetypes=file_types)
                if file_path:
                    if label == "Members skill set":
                        self.skillset_file = file_path
                    else:
                        self.task_details_file = file_path
                else:
                    messagebox.showerror("Error", "No file is selected.")

                entry.insert(tk.END, file_path)
                entry.config(state=tk.DISABLED)
            else:
                # Select a folder
                entry.config(state=tk.NORMAL)
                entry.delete(1.0, tk.END)
                folder_path = filedialog.askdirectory()

                if folder_path:
                    self.ss_document_folder = folder_path
                    # List all file paths in the selected folder and sub-folders
                    folder_file_paths = []
                    for root, _, files in os.walk(folder_path):
                        for file in files:
                            if file.endswith('.xlsx') or file.endswith('.xls'):
                                folder_file_paths.append(os.path.join(root, file))

                    if len(folder_file_paths) > 50:
                        messagebox.showerror("Error", config.error_message["ManyExcelError"])
                        self.ss_folder_file.clear()
                        return
                    else:
                        print("Files in the selected folder and sub-folders:")
                        for file_path in folder_file_paths:
                            entry.insert(tk.END, file_path + "\n")
                            self.ss_folder_file.append(file_path)
                    entry.config(state=tk.DISABLED)
                else:
                    messagebox.showerror("Error", config.error_message["FolderNotFoundError"])
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while selecting the files: {e}")

    def open_url(self, url):
        webbrowser.open_new(url)

    def validate_api_key(self, api_key: str) -> str | None:
        pattern = r'^[A-Za-z0-9]{48}$'
        if not api_key:
            messagebox.showerror("Error", config.error_message["APIEmptyField"])
            return None
        elif re.search(r'[\uFF01-\uFF60\uFFE0-\uFFE6]', api_key):
            messagebox.showerror("Error", config.error_message["FullWidthCharacterError"])
            return None
        elif re.match(pattern, api_key):
            return api_key
        else:
            messagebox.showerror("Error", config.error_message["InvalidKeyError"])
            return None

    # ------------------------------------------------------------------ #
    #  Button state helpers                                               #
    # ------------------------------------------------------------------ #

    def _disable_buttons(self):
        self.btn_start["state"] = tk.DISABLED
        self.btn_skillset["state"] = tk.DISABLED
        self.btn_ss_documents["state"] = tk.DISABLED

    def _enable_buttons(self):
        self.btn_start["state"] = tk.NORMAL
        self.btn_skillset["state"] = tk.NORMAL
        self.btn_ss_documents["state"] = tk.NORMAL

    # ------------------------------------------------------------------ #
    #  Orchestration — main pipeline                                      #
    # ------------------------------------------------------------------ #

    def button_starter(self):
        t = Thread(target=self.main)
        t.start()

    def main(self):
        # if repeat the process, clear the result section first
        self.remove_result_section()

        self.api_key = self.validate_api_key(self.api_key_entry.get())
        if not self.api_key:
            return False

        self._disable_buttons()

        # Read skillset data
        self.skill_set_data = self.read_file(self.skillset_file, 3, None)
        if self.skill_set_data is None:
            return  # Stop the process if the file is not found or an error occurred

        # read and return list of task
        tasks_json = self.read_ss_folder_files()

        # send first request to get complexity and priority
        self.request_task_details(tasks_json)

        # send second request to get wbs details
        self.send_data_to_chatai()

        self._enable_buttons()

    # ------------------------------------------------------------------ #
    #  File reading — delegates to file_parser                            #
    # ------------------------------------------------------------------ #

    def read_file(self, file: str, rowskip: int = 0, rangecol: list | None = None) -> pd.DataFrame | None:
        try:
            return read_excel_file(file, rowskip, rangecol)
        except ValueError:
            messagebox.showerror(
                "Error",
                "The member skillset file is too large. "
                "Please upload a file smaller than 25MB and try again",
            )
            self.browse_file(self.skillset_entry, "Members skill set")
            self._enable_buttons()
            return None
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except pd.errors.EmptyDataError:
            messagebox.showerror("Error", config.error_message["EmptyDataError"])
            return None
        except pd.errors.ParserError:
            messagebox.showerror("Error", config.error_message["ParserError"])
            return None
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
            return None

    # ------------------------------------------------------------------ #
    #  SS folder processing — delegates to file_parser                    #
    # ------------------------------------------------------------------ #

    def read_ss_folder_files(self) -> str:
        self.is_file_valid = True
        self.selected_folder_file = []
        self.task_list.clear()
        is_ss_doc = False

        # Call the method to create the status section
        self.create_result_section()

        self.progress["value"] = 0
        self.status_label.config(text="Processing input data...")
        self.master.after(100, self.process_step)

        # Variables to store file paths based on keywords
        application_detailed_specification_files = []
        screen_layout_files = []

        for file_path in self.ss_folder_file:
            if "Application Detailed Specification" in file_path:
                is_ss_doc = True
                application_detailed_specification_files.append(file_path)
            elif "Screen Layout" in file_path:
                is_ss_doc = True
                screen_layout_files.append(file_path)

        if is_ss_doc == False:
            messagebox.showerror("Error", "No SS documents found in the folder. Please choose the correct folder and try again")
            self.browse_file(self.input_details_entry, "SS Documents")
            self._enable_buttons()
            self.remove_result_section()
            return None

        # Process Screen Layout files
        print("\nScreen Layout Files:")
        for file in screen_layout_files:
            sheetName = "項目定義"
            keywordsHeader = ['画面項目名\n/Screen Item Name', 'タイプ\n/ Type']

            # Validate file
            workbook = load_workbook(filename=file)
            sheet_names = [sheet.title for sheet in workbook.worksheets]
            is_valid, error_message = check_file_validity(sheet_names, workbook, file)

            if not is_valid:
                self.is_file_valid = False

            if self.is_file_valid:
                try:
                    json_str, screen_name = parse_screen_layout(file, sheetName, keywordsHeader)
                    self.screen_layout_json = json_str
                    self.task_list.append(screen_name)
                    print(f"{self.screen_layout_json}")
                except FileNotFoundError:
                    messagebox.showerror("Error", config.error_message["FileNotFoundError"])
                    return None
                except ValueError:
                    messagebox.showerror("Error", config.error_message["EmptyDataError"])
                    return None
                except Exception as e:
                    messagebox.showerror("Error", f"An unexpected error occurred: {e}")
                    return None
            else:
                messagebox.showerror("Error", f"{error_message}")
                self.browse_file(self.input_details_entry, "SS Documents")
                self._enable_buttons()
                self.remove_result_section()
                return None

        # Process Application Detailed Specification files
        print("Application Detailed Specification Files:")
        for file in application_detailed_specification_files:
            workbook = load_workbook(filename=file, data_only=True)
            sheet_names = [sheet.title for sheet in workbook.worksheets]
            is_valid, error_message = check_file_validity(sheet_names, workbook, file)

            if not is_valid:
                self.is_file_valid = False

            if self.is_file_valid:
                try:
                    app_detailed_spec_data = parse_app_detailed_spec(file)
                    json_str, task_names = convert_spec_to_json(app_detailed_spec_data, file)
                    self.app_detailed_spec_data_converted_json = json_str
                    self.task_list.extend(task_names)
                except FileNotFoundError:
                    messagebox.showerror("Error", config.error_message["FileNotFoundError"])
                    return None
                except pd.errors.EmptyDataError:
                    messagebox.showerror("Error", config.error_message["EmptyDataError"])
                    return None
                except pd.errors.ParserError:
                    messagebox.showerror("Error", config.error_message["ParserError"])
                    return None
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to read Excel file: {e}")
                    return None
            else:
                messagebox.showerror("Error", f"{error_message}")
                self.browse_file(self.input_details_entry, "SS Documents")
                self._enable_buttons()
                self.remove_result_section()
                return None

        # Create a list of tasks as JSON
        json_list = [{'Item No': i + 1, 'Task Description': task} for i, task in enumerate(self.task_list)]
        json_string = json.dumps(json_list, indent=4)
        self.progress["value"] = 100

        print(json_string)
        return json_string

    # ------------------------------------------------------------------ #
    #  API requests — delegates to api_client                             #
    # ------------------------------------------------------------------ #

    def request_task_details(self, tasks_list_json: str) -> None:
        try:
            self.progress["value"] = 0
            self.status_label.config(text="Sending request to get task complexity...")
            self.master.after(100, self.process_step)

            prompt = config.prompt_list_task.format(
                screen_layout_json=self.screen_layout_json,
                app_detailed_spec_data_converted_json=self.app_detailed_spec_data_converted_json,
                tasks_list_json=tasks_list_json,
            )

            content = send_gemini_request(self.api_key, prompt)
            self.task_details_response = content
            print("Response from chat AI for the Tasks Complexity")
            print(content)

        except requests.exceptions.RequestException as e:
            if "Too Large" in str(e):
                messagebox.showerror("Error", config.error_message["FileTooBig"])
            else:
                messagebox.showerror("Error", f"Failed to send data to ChatAI: {e}")

        except ValueError as ve:
            print(ve)
            messagebox.showerror("Error", str(ve))
        finally:
            self.progress["value"] = 100

    def send_data_to_chatai(self) -> None:
        start_date_str = self.start_date_entry.get_date()
        end_date_str = self.end_date_entry.get_date()
        start_date = start_date_str.strftime('%m/%d/%Y')
        end_date = end_date_str.strftime('%m/%d/%Y')

        try:
            task_details_data = self.task_details_response
            self.progress["value"] = 0
            self.status_label.config(text="Process Task Details has completed successfully. Sending request to get the WBS details...")
            self.master.after(100, self.process_step)

            prompt = config.prompt.format(
                task_details_data=task_details_data,
                skill_set_data=self.skill_set_data.to_json(),
                start_date_str=start_date,
                end_date_str=end_date,
                task_description="Task Description Example",
                assigned_to="Assigned to Example",
                progress="To do",
                plan_start_date="Start date Example",
                plan_end_date="End date Example",
            )

            content = send_gemini_request(self.api_key, prompt)
            self.create_wbs(content, self.start_date_entry, self.end_date_entry)
            print(content)

        except requests.exceptions.RequestException as e:
            if "Too Large" in str(e):
                messagebox.showerror("Error", config.error_message["FileTooBig"])
            else:
                messagebox.showerror("Error", f"Failed to send data to ChatAI: {e}")

        except ValueError as ve:
            print(ve)
            messagebox.showerror("Error", str(ve))
        finally:
            self.status_label.config(text="Process has completed successfully. You may download the WBS file using the download button below.")
            self.progress["value"] = 100
            self.download_button.grid()
            if hasattr(self, 'progress'):
                self.progress.grid_forget()

    # ------------------------------------------------------------------ #
    #  WBS generation — delegates to wbs_writer                           #
    # ------------------------------------------------------------------ #

    def create_wbs(self, content: str, start_date, end_date) -> None:
        start_date_value = start_date.get_date()
        end_date_value = end_date.get_date()

        try:
            df = markdown_table_to_dataframe(content)
            write_wbs_to_excel(df, start_date_value, end_date_value)
        except FileNotFoundError:
            messagebox.showerror("Error", config.error_message["FileNotFoundError"])
            return None
        except PermissionError as e:
            messagebox.showerror("Error", f"Error: Permission denied. {e}")
            return None
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            return None

    def download_result(self) -> None:
        try:
            destination = copy_to_downloads()
            messagebox.showinfo("Success", f"File saved successfully to {destination}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file: {e}")


